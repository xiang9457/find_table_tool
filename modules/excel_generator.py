#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 生成模块
根据提取的表信息生成 Excel 文件
"""

import openpyxl
from openpyxl.styles import Alignment

class ExcelGenerator:
    """Excel 生成器"""
    
    def __init__(self, excel_path):
        """
        初始化 Excel 生成器
        :param excel_path: Excel 文件路径
        """
        self.excel_path = excel_path
        
    
    def generate(self, table_info_list):
        """
        生成 Excel 文件
        :param table_info_list: 表信息列表
        """
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 创建 Sheet1: 原始表信息
        self._create_sheet1(wb, table_info_list)
        
        # 创建 Sheet2: 清洗后的表信息
        cleaned_table_info = self._clean_table_info(table_info_list)
        self._create_sheet2(wb, cleaned_table_info)
        
        # 创建 Sheet3: 去重后的表信息
        deduplicated_table_info = self._deduplicate_table_info(cleaned_table_info)
        self._create_sheet3(wb, deduplicated_table_info)
        
        # 统计每个 schema 的表数量
        schema_counts = {}
        for table_info in deduplicated_table_info:
            schema = table_info.get('schema', 'master')
            if schema in schema_counts:
                schema_counts[schema] += 1
            else:
                schema_counts[schema] = 1
        
        # 创建 Sheet4: 文件统计信息
        self._create_sheet4(wb, table_info_list)
        
        # 创建 Sheet5: 处理总结
        self._create_sheet5(wb, table_info_list, cleaned_table_info, deduplicated_table_info, schema_counts)
        
        try:
            # 保存 Excel 文件
            wb.save(self.excel_path)
            
            # 打印每个 sheet 的处理结论
            print("   Excel 生成总结:")
            print(f"   - Sheet1 (原始表信息): 共 {len(table_info_list)} 条记录")
            print(f"   - Sheet2 (清洗后表信息): 共 {len(cleaned_table_info)} 条记录，清洗掉 {len(table_info_list) - len(cleaned_table_info)} 条无效记录")
            print(f"   - Sheet3 (去重后表信息): 共 {len(deduplicated_table_info)} 条记录，去重掉 {len(cleaned_table_info) - len(deduplicated_table_info)} 条重复记录")
            print(f"   - Sheet4 (文件统计信息): 已创建")
            print(f"   - Sheet5 (处理总结): 已创建")
            
            print("   - 去重后各 schema 表数量:")
            for schema, count in schema_counts.items():
                print(f"     * {schema}: {count} 个表")
            
            # 打印生成路径信息
            print(f"   Excel 文件生成完成，路径: {self.excel_path}")
        except Exception as e:
            print(f"   生成 Excel 文件时出错: {e}")
            print(f"   错误类型: {type(e).__name__}")
            print(f"   请检查以下情况:")
            print(f"   1. 文件是否被其他程序占用")
            print(f"   2. 磁盘空间是否充足")
            print(f"   3. 是否有写入权限")
            raise
    
    def _create_sheet1(self, wb, table_info_list):
        """
        创建 Sheet1: 原始表信息
        :param wb: 工作簿
        :param table_info_list: 表信息列表
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws1 = wb.active
        ws1.title = "原始表信息"
        
        # 对原始表信息进行排序：从第一列到最后一列升序
        sorted_table_info = sorted(table_info_list, key=lambda x: (
            x.get('source', ''),
            x.get('schema', ''),
            x.get('table_name', ''),
            x.get('file_name', ''),
            x.get('line_num', 0)
        ))
        
        # 设置表头
        headers = ["来源", "schema", "表名", "来源文件名称", "表名所在行号"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        for row, table_info in enumerate(sorted_table_info, 2):
            ws1.cell(row=row, column=1, value=table_info.get('source', ''))
            ws1.cell(row=row, column=2, value=table_info.get('schema', ''))
            ws1.cell(row=row, column=3, value=table_info.get('table_name', ''))
            ws1.cell(row=row, column=4, value=table_info.get('file_name', ''))
            ws1.cell(row=row, column=5, value=table_info.get('line_num', ''))
        
        # 自适应列宽
        for col in range(1, 6):
            max_length = 0
            for row in range(1, ws1.max_row + 1):
                cell = ws1.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2
    
    def _clean_table_info(self, table_info_list):
        """
        清洗表信息
        :param table_info_list: 表信息列表
        :return: 清洗后的表信息列表
        """
        cleaned_table_info = []
        
        # 定义 Java 和 SQL 关键字列表
        keywords = {
            # SQL 关键字
            'SELECT', 'FROM', 'WHERE', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER',
            'TABLE', 'VIEW', 'INDEX', 'TRIGGER', 'PROCEDURE', 'FUNCTION', 'DATABASE', 'SCHEMA',
            'JOIN', 'LEFT', 'RIGHT', 'INNER', 'OUTER', 'ON', 'GROUP', 'BY', 'HAVING', 'ORDER',
            'LIMIT', 'OFFSET', 'AS', 'AND', 'OR', 'NOT', 'IN', 'LIKE', 'BETWEEN', 'IS', 'NULL',
            'TRUE', 'FALSE', 'DISTINCT', 'UNION', 'ALL', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END',
            # Oracle 关键字
            'DUAL', 'TO',
            # Java 关键字
            'public', 'private', 'protected', 'class', 'interface', 'extends', 'implements',
            'static', 'final', 'abstract', 'synchronized', 'volatile', 'transient', 'native',
            'package', 'import', 'if', 'else', 'for', 'while', 'do', 'switch', 'case', 'default',
            'break', 'continue', 'return', 'try', 'catch', 'finally', 'throw', 'throws',
            'new', 'this', 'super', 'instanceof', 'typeof', 'void', 'int', 'long', 'float',
            'double', 'char', 'boolean', 'byte', 'short', 'String', 'Object', 'List', 'Map',
            'Set', 'Array', 'ArrayList', 'HashMap', 'HashSet'
        }
        
        for table_info in table_info_list:
            table_name = table_info.get('table_name', '')
            # 检查是否包含中文
            has_chinese = any('\u4e00' <= c <= '\u9fff' for c in table_name)
            # 检查是否是关键字
            is_keyword = table_name.upper() in keywords or table_name.lower() in keywords
            # 仅保留英文、下划线和点
            cleaned_table_name = ''.join(c for c in table_name if c.isalnum() or c in ['_', '.'])
            # 检查是否包含变量形式
            has_variable = '${' in table_name
            # 排除空表名、包含特殊字符的行、包含中文的行和关键字行
            if cleaned_table_name and not has_variable and not has_chinese and not is_keyword:
                cleaned_table_info.append({
                    'source': table_info.get('source', ''),
                    'schema': table_info.get('schema', ''),
                    'table_name': cleaned_table_name,
                    'file_name': table_info.get('file_name', ''),
                    'line_num': table_info.get('line_num', '')
                })
        
        # 排序：从第一列到最后一列升序
        cleaned_table_info.sort(key=lambda x: (
            x.get('source', ''),
            x.get('schema', ''),
            x.get('table_name', ''),
            x.get('file_name', ''),
            x.get('line_num', 0)
        ))
        
        return cleaned_table_info
    
    def _create_sheet2(self, wb, cleaned_table_info):
        """
        创建 Sheet2: 清洗后的表信息
        :param wb: 工作簿
        :param cleaned_table_info: 清洗后的表信息列表
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws2 = wb.create_sheet(title="清洗后表信息")
        
        # 设置表头
        headers = ["来源", "schema", "表名", "来源文件名称", "表名所在行号"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        for row, table_info in enumerate(cleaned_table_info, 2):
            ws2.cell(row=row, column=1, value=table_info.get('source', ''))
            ws2.cell(row=row, column=2, value=table_info.get('schema', ''))
            ws2.cell(row=row, column=3, value=table_info.get('table_name', ''))
            ws2.cell(row=row, column=4, value=table_info.get('file_name', ''))
            ws2.cell(row=row, column=5, value=table_info.get('line_num', ''))
        
        # 自适应列宽
        for col in range(1, 6):
            max_length = 0
            for row in range(1, ws2.max_row + 1):
                cell = ws2.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2
    
    def _deduplicate_table_info(self, cleaned_table_info):
        """
        去重表信息
        :param cleaned_table_info: 清洗后的表信息列表
        :return: 去重后的表信息列表
        """
        seen = set()
        deduplicated_table_info = []
        
        for table_info in cleaned_table_info:
            key = (table_info.get('schema', ''), table_info.get('table_name', ''))
            if key not in seen:
                seen.add(key)
                deduplicated_table_info.append(table_info)
        
        # 排序：从第一列到最后一列升序
        deduplicated_table_info.sort(key=lambda x: (
            x.get('source', ''),
            x.get('schema', ''),
            x.get('table_name', ''),
            x.get('file_name', ''),
            x.get('line_num', 0)
        ))
        
        return deduplicated_table_info
    
    def _create_sheet3(self, wb, deduplicated_table_info):
        """
        创建 Sheet3: 去重后的表信息
        :param wb: 工作簿
        :param deduplicated_table_info: 去重后的表信息列表
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws3 = wb.create_sheet(title="去重后表信息")
        
        # 设置表头
        headers = ["schema", "表名"]
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 根据schema和表名升序排序
        sorted_table_info = sorted(deduplicated_table_info, key=lambda x: (x.get('schema', ''), x.get('table_name', '')))
        
        # 填充数据
        for row, table_info in enumerate(sorted_table_info, 2):
            ws3.cell(row=row, column=1, value=table_info.get('schema', ''))
            ws3.cell(row=row, column=2, value=table_info.get('table_name', ''))
        
        # 自适应列宽
        max_row = ws3.max_row
        for col in range(1, 3):
            max_length = 0
            for row_num in range(1, max_row + 1):
                cell = ws3.cell(row=row_num, column=col)
                if cell.value:
                    value_str = str(cell.value)
                    max_length = max(max_length, len(value_str))
            column_letter = openpyxl.utils.get_column_letter(col)
            ws3.column_dimensions[column_letter].width = max(max_length + 2, 10)
    
    def _create_sheet4(self, wb, table_info_list):
        """
        创建 Sheet4: 文件统计信息
        :param wb: 工作簿
        :param table_info_list: 表信息列表
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws4 = wb.create_sheet(title="文件统计信息")
        
        # 设置表头
        headers = ["文件类型", "文件数量", "提取表数", "平均每文件"]
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 统计文件类型
        file_types = {}
        table_counts = {}
        
        for table_info in table_info_list:
            file_name = table_info.get('file_name', '')
            if file_name.endswith('.java'):
                file_type = 'Java 文件'
            elif file_name.endswith('.xml'):
                file_type = 'XML 文件'
            else:
                file_type = '其他文件'
            
            # 统计文件数量
            if file_type not in file_types:
                file_types[file_type] = set()
            file_types[file_type].add(file_name)
            
            # 统计表数量
            if file_type not in table_counts:
                table_counts[file_type] = 0
            table_counts[file_type] += 1
        
        # 填充数据
        row = 2
        for file_type in ['Java 文件', 'XML 文件', '其他文件']:
            file_count = len(file_types.get(file_type, set()))
            table_count = table_counts.get(file_type, 0)
            avg_per_file = table_count / file_count if file_count > 0 else 0
            
            ws4.cell(row=row, column=1, value=file_type)
            ws4.cell(row=row, column=2, value=file_count)
            ws4.cell(row=row, column=3, value=table_count)
            ws4.cell(row=row, column=4, value=round(avg_per_file, 2))
            row += 1
        
        # 自适应列宽
        max_row = ws4.max_row
        for col in range(1, 5):
            max_length = 0
            for row_num in range(1, max_row + 1):
                cell = ws4.cell(row=row_num, column=col)
                if cell.value:
                    # 确保计算正确的字符串长度
                    value_str = str(cell.value)
                    max_length = max(max_length, len(value_str))
            # 设置列宽，确保至少有一个合理的最小值
            column_letter = openpyxl.utils.get_column_letter(col)
            ws4.column_dimensions[column_letter].width = max(max_length + 2, 10)
    
    def _create_sheet5(self, wb, table_info_list, cleaned_table_info, deduplicated_table_info, schema_counts):
        """
        创建 Sheet5: 处理总结
        :param wb: 工作簿
        :param table_info_list: 原始表信息列表
        :param cleaned_table_info: 清洗后的表信息列表
        :param deduplicated_table_info: 去重后的表信息列表
        :param schema_counts: Schema 统计信息
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws5 = wb.create_sheet(title="处理总结")
        
        # 设置表头
        cell = ws5.cell(row=1, column=1, value="项目信息")
        # 设置表头样式
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充项目信息
        ws5.cell(row=2, column=1, value="原始记录数")
        ws5.cell(row=2, column=2, value=len(table_info_list))
        
        ws5.cell(row=3, column=1, value="清洗后记录数")
        ws5.cell(row=3, column=2, value=len(cleaned_table_info))
        
        ws5.cell(row=4, column=1, value="清洗掉的记录数")
        ws5.cell(row=4, column=2, value=len(table_info_list) - len(cleaned_table_info))
        
        ws5.cell(row=5, column=1, value="去重后记录数")
        ws5.cell(row=5, column=2, value=len(deduplicated_table_info))
        
        ws5.cell(row=6, column=1, value="去重掉的记录数")
        ws5.cell(row=6, column=2, value=len(cleaned_table_info) - len(deduplicated_table_info))
        
        # Schema 统计
        cell = ws5.cell(row=8, column=1, value="Schema 统计")
        # 设置表头样式
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 9
        for schema, count in schema_counts.items():
            ws5.cell(row=row, column=1, value=f"{schema}")
            ws5.cell(row=row, column=2, value=count)
            row += 1
        
        # 自适应列宽
        max_row = ws5.max_row
        for col in range(1, 3):
            max_length = 0
            for row_num in range(1, max_row + 1):
                cell = ws5.cell(row=row_num, column=col)
                if cell.value:
                    # 确保计算正确的字符串长度
                    value_str = str(cell.value)
                    max_length = max(max_length, len(value_str))
            # 设置列宽，确保至少有一个合理的最小值
            column_letter = openpyxl.utils.get_column_letter(col)
            ws5.column_dimensions[column_letter].width = max(max_length + 2, 10)
